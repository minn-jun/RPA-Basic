from openpyxl import load_workbook

# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져옴
# for row in ws.values:
#     for cell in row:
#         print(cell)


wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 그대로 가져옴
# 12_formula.py 작성한 수식에 대해서는 그대로 들어갈 뿐, 실제로 계산이 안됨.
# evaluate 되지 않은 상태의 데이터는 None 이라고 표시
# 따라서 해당 xlsx 파일을 열었다가 다시 저장하면 계산된 데이터를 받아올 수 있음.
for row in ws.values:
    for cell in row:
        print(cell)