from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학

# move_range("이동범위", rows=이동 줄 수, cols=이동 열 수)
ws.move_range("B1:C11", rows=0, cols=1)
ws["B1"].value = "국어"    # B1셀에 "국어" 입력

# ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("sample_korean.xlsx")